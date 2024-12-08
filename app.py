from flask import Flask, render_template, request, redirect, session
import pymysql.cursors
import re
import os
from openpyxl import load_workbook, Workbook
from datetime import datetime, timedelta

app = Flask(__name__)

# Generate a random secret key
app.secret_key = os.urandom(24)

# MySQL configurations
MYSQL_HOST = 'localhost'
MYSQL_USER = 'root'
MYSQL_PASSWORD = 'Siddhu@123'
MYSQL_DB = 'login'

# File to store login and logout details
LOG_FILE = 'user_activity_logs.xlsx'

def get_db_connection():
    return pymysql.connect(host=MYSQL_HOST,
                           user=MYSQL_USER,
                           password=MYSQL_PASSWORD,
                           db=MYSQL_DB,
                           charset='utf8mb4',
                           cursorclass=pymysql.cursors.DictCursor)

def validate_password(password):
    # Password should contain at least one capital alphabet, one number, one special character, and be more than 6 characters long
    if len(password) < 6:
        return False
    if not re.search("[A-Z]", password):
        return False
    if not re.search("[0-9]", password):
        return False
    if not re.search("[!@#$%^&*()-_=+{};:,<.>?]", password):
        return False
    return True

def format_duration(duration):
    # Format duration as hours:minutes:seconds
    hours, remainder = divmod(duration.total_seconds(), 3600)
    minutes, seconds = divmod(remainder, 60)
    return f"{int(hours):02d}:{int(minutes):02d}:{int(seconds):02d}"

def log_activity(email, activity_type):
    # Log activity (login/logout) time and duration in Excel file
    if not os.path.exists(LOG_FILE):
        wb = Workbook()
        sheet = wb.active
        sheet['A1'] = 'Email'
        sheet['B1'] = 'Login Time'
        sheet['C1'] = 'Logout Time'
        sheet['D1'] = 'Duration'
        wb.save(LOG_FILE)
    
    wb = load_workbook(LOG_FILE)
    sheet = wb.active
    next_row = sheet.max_row + 1
    
    if activity_type == 'Login':
        # Log login time
        sheet.cell(row=next_row, column=1).value = email
        sheet.cell(row=next_row, column=2).value = datetime.now()
    elif activity_type == 'Logout':
        # Find the row with the matching email and log logout time
        for row in range(1, next_row):
            if sheet.cell(row=row, column=1).value == email and sheet.cell(row=row, column=3).value is None:
                sheet.cell(row=row, column=3).value = datetime.now()
                # Calculate duration
                login_time = sheet.cell(row=row, column=2).value
                logout_time = sheet.cell(row=row, column=3).value
                duration = logout_time - login_time
                formatted_duration = format_duration(duration)
                sheet.cell(row=row, column=4).value = formatted_duration
                break
    
    wb.save(LOG_FILE)

@app.route('/', methods=['GET', 'POST'])
def index():
    connection = None  # Initialize connection variable to None

    try:
        if request.method == 'POST':
            if 'register' in request.form:
                email = request.form['email']
                username = request.form['username']
                password = request.form['password']

                # Validate email
                if not email.endswith('@gmail.com') and not email.endswith('.edu.in'):
                    return "Email should end with @gmail.com or edu.in"

                # Validate password
                if not validate_password(password):
                    return "Password should contain at least one capital alphabet, one number, one special character, and be more than 6 characters long."

                # Establish connection and register the user
                connection = get_db_connection()
                with connection.cursor() as cursor:
                    sql = "INSERT INTO users (email, username, password) VALUES (%s, %s, %s)"
                    cursor.execute(sql, (email, username, password))
                connection.commit()
                return redirect('/')

            elif 'login' in request.form:
                username_email = request.form['username_email']
                password = request.form['password']

                # Establish connection and check user credentials
                connection = get_db_connection()
                with connection.cursor() as cursor:
                    if '@' in username_email:
                        # Treat it as an email
                        email = username_email
                        sql = "SELECT * FROM users WHERE email = %s AND password = %s"
                        cursor.execute(sql, (email, password))
                    else:
                        # Treat it as a username
                        username = username_email
                        sql = "SELECT * FROM users WHERE username = %s AND password = %s"
                        cursor.execute(sql, (username, password))
                    user = cursor.fetchone()

                if user:
                    session['loggedin'] = True
                    session['email'] = user['email']
                    log_activity(user['email'], 'Login')  # Log login time
                    return redirect('/dashboard')
                else:
                    return "Invalid username or email or password"
                
    except pymysql.MySQLError as e:
        print(f"MySQL error: {e}")
        return "Database error occurred. Please try again later."
    
    finally:
        if connection is not None:
            connection.close()
    
    return render_template('index.html')

@app.route('/dashboard')
def dashboard():
    if 'loggedin' in session:
        email = session['email']
        connection = None  # Initialize connection variable to None

        try:
            # Establish connection and retrieve user information
            connection = get_db_connection()
            with connection.cursor() as cursor:
                sql = "SELECT username FROM users WHERE email = %s"
                cursor.execute(sql, (email,))
                user = cursor.fetchone()

            if user:
                username = user['username']
                return render_template('dashboard.html', username=username)
            else:
                return "User not found"

        except pymysql.MySQLError as e:
            print(f"MySQL error: {e}")
            return "Database error occurred. Please try again later."
        
        finally:
            if connection is not None:
                connection.close()
    
    return redirect('/')

@app.route('/logout')
def logout():
    if 'loggedin' in session:
        log_activity(session['email'], 'Logout')  # Log logout time
        session.clear()
    return redirect('/')

if __name__ == '__main__':
    app.run(debug=True)
